from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from db.models import *
from .aligments import *
from .borders import *
from .fonts import *


def create_cell(ws, cell: str, text: str | float, font: Font, aligment: Alignment,
                fill: PatternFill = None,
                merge=False,
                end_cell="",
                border=False):
    if border:
        if merge:
            ws = borders_cells(ws, start=cell, end=end_cell, multi=True)
        else:
            ws = borders_cells(ws, start=cell)
    if merge:
        ws.merge_cells(f'{cell}:{end_cell}')

    ws[cell] = text
    ws[cell].font = font
    ws[cell].alignment = aligment
    if fill:
        ws[cell].fill = fill

    return ws


def auto_size(ws):
    for column_cells in ws.columns:
        unmerged_cells = list(
            filter(lambda cell_to_check: cell_to_check.coordinate not in ws.merged_cells, column_cells))
        length = max(len(str(cell.value)) for cell in unmerged_cells)
        ws.column_dimensions[unmerged_cells[0].column_letter].width = length * 1.2


def sort_procedures_by_date(p: User_Procedure):
    return p.created_at


def statistics_to_excel_cur_month(data, date: datetime, month_max: int) -> Workbook:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Статистика"
    ws1 = create_cell(ws1, cell="A1", end_cell="A2", text='', aligment=al_center, font=ft_Arial_9,
                      fill=PatternFill('solid', fgColor="1F4E78"), merge=True)

    ws1 = create_cell(ws1,
                      cell="B1",
                      end_cell=f"{get_column_letter(month_max + 2)}1",
                      text=f"Статистика c 1.{date.month}.{date.year} по {month_max}.{date.month}.{date.year}",
                      font=ft_Arial_10_bold,
                      aligment=al_center,
                      fill=PatternFill('solid', fgColor="1F4E78"),
                      border=True,
                      merge=True,
                      )
    for i in range(1, month_max + 1):
        letter = get_column_letter(i + 1)
        ws1 = create_cell(ws1,
                          cell=f'{letter}2',
                          text=f'{i}.{f"0{date.month}" if date.month < 10 else date.month}',
                          font=ft_Arial_9_bold,
                          aligment=al_center,
                          border=True,
                          fill=PatternFill('solid', fgColor="4F93D1"),
                          )
    ws1 = create_cell(ws1,
                      cell=f"{get_column_letter(month_max + 2)}2",
                      text="Итог за месяц",
                      font=ft_Arial_9_bold,
                      aligment=al_center,
                      border=True,
                      fill=PatternFill('solid', fgColor="4F93D1"),
                      )

    start_cell = 3
    change_color = True
    all_day_money = {datetime(year=date.year, month=date.month, day=i).date(): Decimal('0') for i in
                     range(1, month_max + 1)}

    for brigade in data:
        brigade_row = start_cell
        brigade_day_money = {datetime(year=date.year, month=date.month, day=i).date(): Decimal('0') for i in
                             range(1, month_max + 1)}
        brigade_month = Decimal('0')
        ws1 = create_cell(ws1,
                          cell=f"A{brigade_row}",
                          text=f"Бригада №{brigade.number}({brigade.name})",
                          font=ft_Arial_9_bold,
                          aligment=al_wrap,
                          border=True,
                          fill=PatternFill('solid', fgColor=f"{'BDD7EE' if change_color else 'DDEBF7'}"))
        start_cell += 1

        for user in brigade.user:
            user_row = start_cell
            ws1 = create_cell(ws1,
                              cell=f"A{user_row}",
                              text=f"{user.name} {user.surname}",
                              font=ft_Arial_9,
                              aligment=al_center,
                              border=True,
                              fill=PatternFill('solid', fgColor=f"{'BDD7EE' if change_color else 'DDEBF7'}")
                              )

            user_month_dict = {datetime(year=date.year, month=date.month, day=i).date(): Decimal('0') for i in
                               range(1, month_max + 1)}
            for proc in sorted(user.procedures, key=sort_procedures_by_date):
                user_month_dict[proc.created_at] += Decimal(proc.count) * proc.procedure.rate

            user_month = Decimal('0')
            for i, key in enumerate(user_month_dict.keys()):
                ws1 = create_cell(ws1,
                                  cell=f"{get_column_letter(i + 2)}{user_row}",
                                  text=f"{user_month_dict[key]}",
                                  font=ft_Arial_9,
                                  aligment=al_center,
                                  fill=PatternFill('solid', fgColor=f"{'BDD7EE' if change_color else 'DDEBF7'}"),
                                  )
                brigade_day_money[key] += user_month_dict[key]
                all_day_money[key] += user_month_dict[key]
                user_month += user_month_dict[key]
            ws1 = create_cell(ws1,
                              cell=f"{get_column_letter(month_max + 2)}{user_row}",
                              text=f"{user_month}",
                              font=ft_Arial_9,
                              aligment=al_center,
                              fill=PatternFill('solid', fgColor="9BC2E6"),
                              )
            start_cell += 1
        for i, key in enumerate(brigade_day_money.keys()):
            ws1 = create_cell(ws1,
                              cell=f"{get_column_letter(i + 2)}{brigade_row}",
                              text=f"{brigade_day_money[key]}",
                              font=ft_Arial_9_bold,
                              aligment=al_center,
                              fill=PatternFill('solid', fgColor=f"{'BDD7EE' if change_color else 'DDEBF7'}"),
                              )
            brigade_month += brigade_day_money[key]
        ws1 = create_cell(ws1,
                          cell=f"{get_column_letter(month_max + 2)}{brigade_row}",
                          text=f"{brigade_month}",
                          font=ft_Arial_9_bold,
                          aligment=al_center,
                          fill=PatternFill('solid', fgColor="9BC2E6")
                          )
        change_color = not change_color

    ws1 = create_cell(ws1,
                      cell=f"A{start_cell}",
                      text=f"Итого за день",
                      font=ft_Arial_9_bold,
                      aligment=al_center,
                      border=True,
                      fill=PatternFill('solid', fgColor="9BC2E6"),
                      )

    all_money = Decimal('0')
    for i, key in enumerate(all_day_money.keys()):
        ws1 = create_cell(ws1,
                          cell=f"{get_column_letter(i + 2)}{start_cell}",
                          text=f"{all_day_money[key]}",
                          font=ft_Arial_9_bold,
                          aligment=al_center,
                          fill=PatternFill('solid', fgColor="9BC2E6"),
                          )
        all_money += all_day_money[key]
    ws1 = create_cell(ws1,
                      cell=f"{get_column_letter(month_max + 2)}{start_cell}",
                      text=f"{all_money}",
                      font=ft_Arial_9_bold,
                      aligment=al_center,
                      fill=PatternFill('solid', fgColor="9BC2E6"),
                      )

    ws1 = create_cell(ws1,
                      cell=f"A{start_cell + 4}",
                      end_cell=f"C{start_cell + 8}",
                      text=f"Если по какой-то причине вы не можете найти сотрудника в списке," +
                           " вероятнее всего он/она не сделали ни одной операции за месяц, " +
                           "либо что-то не так с программой, в таком случае просьба написать на web@7versts.ru",
                      font=ft_Arial_9,
                      aligment=al_wrap,
                      border=True,
                      merge=True,
                      )

    auto_size(ws1)
    return wb
